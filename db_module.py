import psycopg2
import psycopg2.extras
from psycopg2.sql import SQL, Identifier

import datetime 

import openpyxl
from openpyxl.styles import Font

class DB_module:
    
    def __init__(self, DB_HOST, DB_NAME, DB_USER, DB_PASS, DB_PORT):
        self.DB_HOST = DB_HOST
        self.DB_NAME = DB_NAME
        self.DB_USER = DB_USER
        self.DB_PASS = DB_PASS
        self.DB_PORT = DB_PORT
         
    def get_answer(self, answer_id = None, user_id = None, answering_date = None):
    
        with psycopg2.connect(
                host = self.DB_HOST,
                database = self.DB_NAME,
                user = self.DB_USER,
                password = self.DB_PASS,
                port = self.DB_PORT
                ) as conn:
            
            with conn.cursor(cursor_factory=psycopg2.extras.DictCursor) as cur:
                                
                if (answer_id != None):
                    cur.execute(f"select * from answer_loyalty where id = {answer_id};")
                    return cur.fetchone()
                elif (user_id != None):
                    cur.execute(f"select * from answer_loyalty where user_id = {user_id};")
                    lines = cur.fetchall()
                    max_id = 0
                    for ans in lines: max_id = max(int(ans['id']), max_id)
                    return self.get_answer(answer_id = max_id)
                elif (answering_date != None):
                    cur.execute(f"select * from answer_loyalty where answering_date = {answering_date};")
                    return cur.fetchall()
                else:
                    cur.execute("select * from answer_loyalty")
                    return cur.fetchall()
    
    
    
    def add_answer(self, user_id = None, question_name = None, answer = None):
        
        first_question_name = 'loyalty'
    
        with psycopg2.connect(
                host = self.DB_HOST,
                database = self.DB_NAME,
                user = self.DB_USER,
                password = self.DB_PASS,
                port = self.DB_PORT
                ) as conn:
            
            with conn.cursor(cursor_factory=psycopg2.extras.DictCursor) as cur:
                
                if question_name != first_question_name: 
                    cur.execute(f"select * from answer_loyalty where user_id = {user_id};")
                    lines = cur.fetchall()
                    max_id = 0
                    for ans in lines: max_id = max(int(ans['id']), max_id)
                    
                    cur.execute(SQL("""
                                UPDATE answer_loyalty SET 
                                {} = %(answer)s
                                WHERE
                                id = %(id)s
                                ;
                                """).format(Identifier(question_name)), 
                                {
                                'id': max_id,
                                'answer': answer
                                })
                    
                
                else: # 
                    answering_date = datetime.date.today()
                    cur.execute("""
                                insert into answer_loyalty
                                (user_id, answering_date, loyalty, review) 
                                values
                                (%(user_id)s, %(answering_date)s, %(loyalty)s, '-');
                                """, 
                                {
                                 'user_id': user_id, 
                                 'answering_date': answering_date, 
                                 'loyalty': answer,
                                 })
                return 
            
    def add_review(self, user_id = None, text = None):
    
        with psycopg2.connect(
                host = self.DB_HOST,
                database = self.DB_NAME,
                user = self.DB_USER,
                password = self.DB_PASS,
                port = self.DB_PORT
                ) as conn:
            
            with conn.cursor(cursor_factory=psycopg2.extras.DictCursor) as cur:
                
                cur.execute(f"select * from answer_loyalty where user_id = {user_id};")
                lines = cur.fetchall()
                max_id = 0
                for ans in lines: max_id = max(int(ans['id']), max_id)
                
                cur.execute("""
                                UPDATE answer_loyalty SET 
                                review = review || %(text)s
                                WHERE
                                id = %(id)s
                                ;
                                """, 
                                {
                                'id': max_id,
                                'text': text+'\n'
                                })
                                
                                
                                
    def get_user(self, user_id):
        
        with psycopg2.connect(
                host = self.DB_HOST,
                database = self.DB_NAME,
                user = self.DB_USER,
                password = self.DB_PASS,
                port = self.DB_PORT
                ) as conn:
            
            with conn.cursor(cursor_factory=psycopg2.extras.DictCursor) as cur:
                
                cur.execute(f"select * from users where id = {user_id};")
                return cur.fetchone()
                
                
        
            
    def add_user(self, user_id, first_name, last_name, username, language_code):
        
           
        with psycopg2.connect(
                host = self.DB_HOST,
                database = self.DB_NAME,
                user = self.DB_USER,
                password = self.DB_PASS,
                port = self.DB_PORT
                ) as conn:
            
            with conn.cursor(cursor_factory=psycopg2.extras.DictCursor) as cur:
                
                if self.get_user(user_id): return 'user exist'
                
                cur.execute("""
                            insert into users
                            (id, first_name, last_name, username, language_code) 
                            values
                            (%(id)s, %(first_name)s, %(last_name)s, %(username)s, %(language_code)s);
                            """, 
                            {
                             'id': user_id, 
                             'first_name': first_name, 
                             'last_name': last_name,
                             'username': username,
                             'language_code': language_code
                             })
                return 'user added'
            
    def add_phone(self, user_id = None, phone = None):
    
        with psycopg2.connect(
                host = self.DB_HOST,
                database = self.DB_NAME,
                user = self.DB_USER,
                password = self.DB_PASS,
                port = self.DB_PORT
                ) as conn:
            
            with conn.cursor(cursor_factory=psycopg2.extras.DictCursor) as cur:
                
                cur.execute("""
                                UPDATE users SET 
                                phone = %(phone)s
                                WHERE
                                id = %(id)s
                                ;
                                """, 
                                {
                                'id': user_id,
                                'phone': phone
                                })
    
    def export_to_excel(self, headings, filepath):

        """
        Exports data from PostgreSQL to an Excel spreadsheet using psycopg2.
    
        Arguments:
        connection - an open psycopg2 (this function does not close the connection)
        query_string - SQL to get data
        headings - list of strings to use as column headings
        filepath - path and filename of the Excel file
    
        psycopg2 and file handling errors bubble up to calling code.
        """
        
        with psycopg2.connect(
                host = self.DB_HOST,
                database = self.DB_NAME,
                user = self.DB_USER,
                password = self.DB_PASS,
                port = self.DB_PORT
                ) as conn:
            
            with conn.cursor(cursor_factory=psycopg2.extras.DictCursor) as cur:
                cur.execute("""
                            select u.*, al.* 
                            from 
                            answer_loyalty al 
                            left join 
                            users u 
                            on 
                            u.id = al.user_id;
                            """)
                data = cur.fetchall()
    
    
        wb = openpyxl.Workbook()
        sheet = wb.create_sheet(title='Все ответы', index=0)
    
        sheet.row_dimensions[1].font = Font(bold = True)
    
        # Spreadsheet row and column indexes start at 1
        # so we use "start = 1" in enumerate so
        # we don't need to add 1 to the indexes.
        for colno, heading in enumerate(headings, start = 1):
            sheet.cell(row = 1, column = colno).value = heading
    
        # This time we use "start = 2" to skip the heading row.
        for rowno, row in enumerate(data, start = 2):
            for colno, cell_value in enumerate(row, start = 1):
                sheet.cell(row = rowno, column = colno).value = cell_value
    
        wb.save(filepath) 

        